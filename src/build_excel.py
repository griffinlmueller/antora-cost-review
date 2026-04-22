import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

RED, ORANGE, YELLOW, PURPLE = "FFCDD2", "FFE0B2", "FFF9C4", "E1BEE7"
TEAL, WHITE = "0D7B6E", "FFFFFF"

def build_excel(tracker_df, issues):
    flags = {}
    for e in issues["stale_entries"]:
        flags.setdefault(e["id"], []).append(("STALE", RED))
    for dup in issues["duplicates"]:
        for _id in dup["ids"]:
            flags.setdefault(_id, []).append(("DUPLICATE", ORANGE))
    for c in issues["conflicting_assumptions"]:
        for _id in c["ids"]:
            flags.setdefault(_id, []).append(("CONFLICT", YELLOW))
    for h in issues["high_savings_no_owner"]:
        flags.setdefault(h["id"], []).append(("NO OWNER", PURPLE))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Tracker Review"

    ws.merge_cells("A1:O1")
    ws["A1"] = "ANTORA COST OPPORTUNITY TRACKER — AI REVIEW"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", start_color=TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    legend = [("STALE (6+ mo)", RED), ("DUPLICATE", ORANGE), ("ASSUMPTION CONFLICT", YELLOW), ("HIGH SAVINGS, NO OWNER", PURPLE)]
    for i, (label, color) in enumerate(legend):
        cell = ws.cell(row=3, column=i*3+1, value=label)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center")

    headers = list(tracker_df.columns) + ["flags"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h.replace("_", " ").title())
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=TEAL)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[5].height = 20

    for row_idx, row in tracker_df.iterrows():
        excel_row = row_idx + 6
        row_flags = flags.get(str(row["id"]), [])
        row_color = row_flags[0][1] if row_flags else WHITE

        for col_idx, col_name in enumerate(tracker_df.columns, 1):
            c = ws.cell(row=excel_row, column=col_idx, value=row[col_name])
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
            if row_flags:
                c.fill = PatternFill("solid", start_color=row_color)

        flag_text = " | ".join(f[0] for f in row_flags) if row_flags else ""
        c = ws.cell(row=excel_row, column=len(tracker_df.columns)+1, value=flag_text)
        c.font = Font(name="Arial", size=9, bold=True, color="C62828" if row_flags else "000000")
        if row_flags:
            c.fill = PatternFill("solid", start_color=row_color)

    widths = [6, 24, 22, 40, 28, 14, 14, 14, 14, 14, 14, 14, 36, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf